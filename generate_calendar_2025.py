#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from datetime import datetime, date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import calendar

def is_weekend(dt):
    """判断是否为周末"""
    return dt.weekday() >= 5

def get_2025_holidays():
    """获取2025年中国法定节假日"""
    holidays = {
        # 元旦
        '2025-01-01': '元旦',
        
        # 春节 (农历除夕至正月初七)
        '2025-01-28': '春节',
        '2025-01-29': '春节',
        '2025-01-30': '春节',
        '2025-01-31': '春节',
        '2025-02-01': '春节',
        '2025-02-02': '春节',
        '2025-02-03': '春节',
        '2025-02-04': '春节',
        
        # 清明节
        '2025-04-04': '清明节',
        '2025-04-05': '清明节',
        '2025-04-06': '清明节',
        
        # 劳动节
        '2025-05-01': '劳动节',
        '2025-05-02': '劳动节',
        '2025-05-03': '劳动节',
        '2025-05-04': '劳动节',
        '2025-05-05': '劳动节',
        
        # 端午节
        '2025-05-31': '端午节',
        '2025-06-01': '端午节',
        '2025-06-02': '端午节',
        
        # 中秋节和国庆节
        '2025-10-01': '中秋节、国庆节',
        '2025-10-02': '中秋节、国庆节',
        '2025-10-03': '中秋节、国庆节',
        '2025-10-04': '中秋节、国庆节',
        '2025-10-05': '中秋节、国庆节',
        '2025-10-06': '中秋节、国庆节',
        '2025-10-07': '中秋节、国庆节',
        '2025-10-08': '中秋节、国庆节',
    }
    return holidays

def get_2025_workdays():
    """获取2025年调休工作日（周末需要上班的日期）"""
    workdays = {
        # 春节调休
        '2025-01-26': '春节调休',
        '2025-02-08': '春节调休',
        
        # 劳动节调休
        '2025-04-27': '劳动节调休',
        
        # 中秋节、国庆节调休
        '2025-09-28': '中秋节、国庆节调休',
        '2025-10-11': '中秋节、国庆节调休',
    }
    return workdays

def generate_2025_calendar():
    """生成2025年中国日历"""
    
    holidays = get_2025_holidays()
    workdays = get_2025_workdays()
    
    # 生成全年日期数据
    calendar_data = []
    start_date = date(2025, 1, 1)
    end_date = date(2025, 12, 31)
    
    current_date = start_date
    while current_date <= end_date:
        date_str = current_date.strftime('%Y-%m-%d')
        weekday = current_date.weekday()
        weekday_name = ['周一', '周二', '周三', '周四', '周五', '周六', '周日'][weekday]
        
        # 判断日期类型
        if date_str in holidays:
            date_type = '法定节假日'
            holiday_name = holidays[date_str]
        elif date_str in workdays:
            date_type = '调休工作日'
            holiday_name = workdays[date_str]
        elif is_weekend(current_date):
            date_type = '周末休息'
            holiday_name = ''
        else:
            date_type = '工作日'
            holiday_name = ''
        
        calendar_data.append({
            '日期': date_str,
            '星期': weekday_name,
            '日期类型': date_type,
            '节假日名称': holiday_name
        })
        
        current_date += timedelta(days=1)
    
    return calendar_data

def create_full_calendar_excel(calendar_data):
    """创建包含2025整年每一天的Excel文件"""
    
    # 创建Excel文件
    filename = '日历.xlsx'
    wb = Workbook()
    
    # 删除默认sheet
    if wb.active:
        wb.remove(wb.active)
    
    # 创建sheet
    ws = wb.create_sheet(title='2025年节假日安排')
    
    # 设置表头
    headers = ['日期', '星期', '日期类型', '节假日名称']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=12, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 填充数据
    for row_idx, row_data in enumerate(calendar_data, 2):
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data[header])
            
            # 设置样式
            if row_data['日期类型'] == '法定节假日':
                cell.fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')  # 红色
            elif row_data['日期类型'] == '调休工作日':
                cell.fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')  # 橙色
            elif row_data['日期类型'] == '周末休息':
                cell.fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')  # 浅紫色
            else:  # 工作日
                cell.fill = PatternFill(start_color='F0F8FF', end_color='F0F8FF', fill_type='solid')  # 浅蓝色
            
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    # 设置列宽
    column_widths = [15, 10, 15, 20]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # 保存文件
    wb.save(filename)
    print(f'2025年节假日安排日历已生成: {filename}')

if __name__ == '__main__':
    calendar_data = generate_2025_calendar()
    create_full_calendar_excel(calendar_data) 