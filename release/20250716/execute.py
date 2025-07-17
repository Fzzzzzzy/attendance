import pandas as pd
import numpy as np
from datetime import datetime, timedelta, date
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import re
from typing import List, Tuple, Dict, Set, Optional

def read_excel_file(file_path: str, sheet_name: str) -> Tuple[Optional[pd.DataFrame], str]:
    """通用的Excel文件读取函数，返回DataFrame和错误信息"""
    try:
        if not os.path.exists(file_path):
            return None, f"文件不存在: {file_path}"
        
        xl_file = pd.ExcelFile(file_path)
        sheet_names = xl_file.sheet_names
        
        if sheet_name not in sheet_names:
            return None, f"在文件 {file_path} 中未找到名为 '{sheet_name}' 的sheet"
        
        df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')
        print(f"成功读取 {file_path} 的 {sheet_name} sheet，共 {len(df)} 行数据")
        return df, ""
        
    except Exception as e:
        return None, f"读取文件 {file_path} 时出错: {e}"

def apply_cell_format(cell, value: str, is_header: bool = False, 
                     bg_color: str = None, font_color: str = '000000',
                     bold: bool = False, font_size: int = None):
    """通用的单元格格式设置函数"""
    cell.value = value
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    if bg_color:
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
    
    font_args = {
        'bold': bold,
        'color': font_color
    }
    if font_size:
        font_args['size'] = font_size
    cell.font = Font(**font_args)
    
    if not is_header:
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

def set_column_widths(worksheet, widths: List[int]):
    """设置工作表的列宽"""
    for i, width in enumerate(widths, 1):
        worksheet.column_dimensions[chr(64 + i)].width = width

def set_row_heights(worksheet, start_row: int, end_row: int, height: int):
    """设置工作表的行高"""
    for row in range(start_row, end_row + 1):
        worksheet.row_dimensions[row].height = height

def parse_datetime_str(date_time_str: str, is_end_time: bool = False) -> datetime:
    """解析日期时间字符串为datetime对象

    Args:
        date_time_str (str): 日期时间字符串，格式可能为：
            - '2025-04-28'
            - '2025-04-28 上午'
            - '2025-04-28 下午'
            - '2025-04-28 11:00'
        is_end_time (bool): 是否是结束时间，用于确定默认时间

    Returns:
        datetime: 解析后的datetime对象

    示例：
        - parse_datetime_str('2025-04-28', False) -> 2025-04-28 09:00:00
        - parse_datetime_str('2025-04-28', True) -> 2025-04-28 18:00:00
        - parse_datetime_str('2025-04-28 上午') -> 2025-04-28 09:00:00
        - parse_datetime_str('2025-04-28 下午') -> 2025-04-28 13:00:00
        - parse_datetime_str('2025-04-28 上午', True) -> 2025-04-28 12:00:00
        - parse_datetime_str('2025-04-28 下午', True) -> 2025-04-28 18:00:00
        - parse_datetime_str('2025-04-28 11:00') -> 2025-04-28 11:00:00
    """
    try:
        # 移除可能的多余空格
        date_time_str = date_time_str.strip()
        
        # 尝试直接解析完整的日期时间格式
        try:
            return datetime.strptime(date_time_str, '%Y-%m-%d %H:%M')
        except ValueError:
            pass

        # 分割日期和时间部分
        parts = date_time_str.split()
        date_str = parts[0]  # 日期部分
        time_str = parts[1] if len(parts) > 1 else None  # 时间部分（如果有）

        # 解析日期部分
        base_date = datetime.strptime(date_str, '%Y-%m-%d')

        # 如果没有时间部分，使用默认时间
        if not time_str:
            default_time = "18:00" if is_end_time else "09:00"
            return datetime.strptime(f"{date_str} {default_time}", '%Y-%m-%d %H:%M')

        # 处理特殊时间标记
        if "上午" in time_str:
            if is_end_time:
                # 上午结束时间为12:00
                return datetime.strptime(f"{date_str} 12:00", '%Y-%m-%d %H:%M')
            else:
                # 上午开始时间为09:00
                return datetime.strptime(f"{date_str} 09:00", '%Y-%m-%d %H:%M')
        elif "下午" in time_str:
            if is_end_time:
                # 下午结束时间为18:00
                return datetime.strptime(f"{date_str} 18:00", '%Y-%m-%d %H:%M')
            else:
                # 下午开始时间为13:00
                return datetime.strptime(f"{date_str} 13:00", '%Y-%m-%d %H:%M')
        else:
            # 如果是具体时间，直接使用
            return datetime.strptime(f"{date_str} {time_str}", '%Y-%m-%d %H:%M')

    except Exception as e:
        raise ValueError(f"无法解析日期时间: {date_time_str}, 错误: {str(e)}")

def get_status_color(status: str) -> Optional[str]:
    """根据状态返回对应的颜色代码"""
    status_colors = {
        '出勤不足': 'E6B8AF',  # 莫兰迪粉红
        '未打卡': 'B6D7A8',    # 莫兰迪绿
        '上/下班漏打卡': 'FFE599'  # 莫兰迪黄
    }
    return status_colors.get(status)

def filter_employees(employee_df):
    """筛选符合条件的员工：全职 + Grade < 13 + 工作地点为上海"""
    if employee_df.empty:
        return pd.DataFrame()
    
    # 筛选全职员工
    full_time_mask = employee_df['员工类型'] == '全职'
    full_time_df = employee_df[full_time_mask].copy()
    
    # 筛选工作地点为上海
    shanghai_mask = full_time_df['工作地点'] == '上海'
    shanghai_df = full_time_df[shanghai_mask].copy()
    
    # 筛选Grade < 13的员工
    def extract_grade(grade_str):
        if pd.isna(grade_str) or grade_str == '':
            return 999  # 默认给一个很大的值，确保被过滤掉
        # 提取Grade后面的数字
        match = re.search(r'Grade\s*(\d+)', str(grade_str))
        if match:
            return int(match.group(1))
        return 999
    
    shanghai_df['grade_num'] = shanghai_df['职级'].apply(extract_grade)
    qualified_df = shanghai_df[shanghai_df['grade_num'] < 13].copy()
    qualified_df = qualified_df.drop('grade_num', axis=1)
    
    print(f"筛选出符合条件的员工: {len(qualified_df)} 人")
    return qualified_df

def get_absence_description(name: str, day: str, absence_dfs: List[Tuple[pd.DataFrame, str, str, str]]) -> str:
    """获取请假情况说明

    Args:
        name (str): 员工姓名
        day (str): 日期
        absence_dfs (List[Tuple[pd.DataFrame, str, str, str]]): 请假数据表列表，每个元素为(数据表,类型列,开始时间列,结束时间列)

    Returns:
        str: 请假情况说明，格式为"请假类型[HH:MM-HH:MM]"，多个请假用分号分隔
    """
    all_descriptions = []
    
    # 解析目标日期和工作时间范围
    target_date = datetime.strptime(day, '%Y-%m-%d')
    workday_start = target_date.replace(hour=9, minute=0)  # 当天9:00
    workday_end = target_date.replace(hour=18, minute=0)   # 当天18:00

    # 遍历所有数据表
    for df, type_col, start_col, end_col in absence_dfs:
        if df.empty:
            continue

        # 获取当前员工在当前表中的所有记录
        records = df[
            (df['姓名'] == name) & 
            (df['状态'].isin(['已生效', '未生效']))
        ]
        
        if records.empty:
            continue

        # 处理每条记录
        for _, record in records.iterrows():
            # 使用process_absence_record获取时间段交集
            time_period = process_absence_record(
                record, start_col, end_col, target_date, workday_start, workday_end
            )
            
            if time_period:
                # 获取请假类型
                type_val = str(record[type_col]) if type_col in record and pd.notna(record[type_col]) else ''
                
                # 格式化时间段
                start_time = time_period[0].strftime('%H:%M')
                end_time = time_period[1].strftime('%H:%M')
                
                # 构建描述
                description = f"{type_val}[{start_time}-{end_time}]"
                all_descriptions.append(description)

    return '; '.join(all_descriptions)

def merge_two_datetime_periods(period1: Tuple[datetime, datetime], period2: Tuple[datetime, datetime]) -> Optional[Tuple[datetime, datetime]]:
    """合并两个时间段，如果它们重叠或相邻则返回合并后的时间段，否则返回None

    Args:
        period1 (Tuple[datetime, datetime]): 第一个时间段
        period2 (Tuple[datetime, datetime]): 第二个时间段

    Returns:
        Optional[Tuple[datetime, datetime]]: 合并后的时间段，如果不能合并则返回None
    """
    start1, end1 = period1
    start2, end2 = period2
    
    # 检查是否重叠
    if end1 < start2 or end2 < start1:
        return None
        
    # 如果重叠，返回最早的开始时间和最晚的结束时间
    return (min(start1, start2), max(end1, end2))

def merge_datetime_periods(periods: List[Tuple[datetime, datetime]]) -> List[Tuple[datetime, datetime]]:
    """合并重叠或相邻的时间段

    Args:
        periods (List[Tuple[datetime, datetime]]): 待合并的时间段列表

    Returns:
        List[Tuple[datetime, datetime]]: 合并后的时间段列表
    """
    if not periods:
        return []
    
    # 按开始时间排序
    sorted_periods = sorted(periods, key=lambda x: x[0])

    i = 0
    while i < len(sorted_periods) - 1:
        result = merge_two_datetime_periods(sorted_periods[i], sorted_periods[i + 1])
        if result:
            sorted_periods[i] = result
            sorted_periods.pop(i + 1)
        else:
            i += 1
            
    return sorted_periods

def calculate_datetime_minutes(periods: List[Tuple[datetime, datetime]]) -> int:
    """计算时间段列表的总分钟数

    Args:
        periods (List[Tuple[datetime, datetime]]): 时间段列表

    Returns:
        int: 总分钟数
    """
    total_minutes = 0
    for start, end in periods:
        delta = end - start
        total_minutes += int(delta.total_seconds() / 60)
    return total_minutes

def process_absence_record(record, start_col: str, end_col: str, target_date: datetime, workday_start: datetime, workday_end: datetime) -> Optional[Tuple[datetime, datetime]]:
    """处理单条请假记录

    Args:
        record: 请假记录（DataFrame的一行）
        start_col (str): 开始时间列名
        end_col (str): 结束时间列名
        target_date (datetime): 目标日期
        workday_start (datetime): 工作日开始时间
        workday_end (datetime): 工作日结束时间

    Returns:
        Optional[Tuple[datetime, datetime]]: 时间段
        如果请假记录无效或没有交集，返回 None
        否则返回 (开始时间, 结束时间)
    """
    try:
        # 解析开始和结束时间
        try:
            start_time = parse_datetime_str(str(record[start_col]))
            end_time = parse_datetime_str(str(record[end_col]), is_end_time=True)
        except Exception as e:
            print(f"解析时间出错: {e}")
            return None

        # 如果请假记录与目标日期没有重叠，跳过
        if end_time.date() < target_date.date() or start_time.date() > target_date.date():
            return None

        # 计算请假时间段和工作时间段的交集
        intersection_start = max(start_time, workday_start)
        intersection_end = min(end_time, workday_end)

        # 如果没有交集，跳过此记录
        if intersection_start >= intersection_end:
            return None

        # 返回交集时间段
        return (intersection_start, intersection_end)

    except Exception as e:
        print(f"处理请假记录时出错: {e}")
        return None

def get_raw_absence_hours(name: str, day: str, absence_dfs: List[Tuple[pd.DataFrame, str, str, str]]) -> Tuple[float, List[Tuple[datetime, datetime]]]:
    """获取员工某天的原始请假时长和时间段

    Args:
        name (str): 员工姓名
        day (str): 日期
        absence_dfs (List[Tuple[pd.DataFrame, str, str, str]]): 请假数据表列表，每个元素为(数据表,类型列,开始时间列,结束时间列)

    Returns:
        Tuple[float, List[Tuple[datetime, datetime]]]: (原始请假时长, 合并后的时间段列表)
    """
    all_time_periods = []

    # 解析目标日期和工作时间范围
    target_date = datetime.strptime(day, '%Y-%m-%d')
    workday_start = target_date.replace(hour=9, minute=0)  # 当天9:00
    workday_end = target_date.replace(hour=18, minute=0)   # 当天18:00

    # 遍历所有请假类型的数据表
    for df, type_col, start_col, end_col in absence_dfs:
        # 获取当前员工的请假记录
        records = df[
            (df['姓名'] == name) & 
            (df['状态'].isin(['已生效', '未生效']))
        ]

        # 处理每条请假记录
        for _, record in records.iterrows():
            time_period = process_absence_record(
                record, start_col, end_col, target_date, workday_start, workday_end
            )
            
            if time_period:
                all_time_periods.append(time_period)

    # 如果没有有效的时间段，返回0
    if not all_time_periods:
        return 0.0, []

    # 合并重叠的时间段
    merged_periods = merge_datetime_periods(all_time_periods)

    # 计算总时长（小时）
    total_minutes = calculate_datetime_minutes(merged_periods)
    total_hours = total_minutes / 60

    # 如果总时长超过9小时，返回9小时，但仍然返回实际的时间段
    return min(total_hours, 9.0), merged_periods

def get_actual_absence_hours(merged_periods: List[Tuple[datetime, datetime]]) -> float:
    """计算实际请假时长（考虑午休时间）

    Args:
        merged_periods (List[Tuple[datetime, datetime]]): 合并后的时间段列表

    Returns:
        float: 实际请假时长

    Notes:
        在原有时间段的基础上添加午休时间段(12:00-13:00)，重新合并后计算实际时长
    """
    if not merged_periods:
        return 0.0

    # 获取第一个时间段的日期作为参考
    reference_date = merged_periods[0][0].date()
    
    # 添加午休时间段
    lunch_break_start = datetime.combine(reference_date, datetime.strptime('12:00', '%H:%M').time())
    lunch_break_end = datetime.combine(reference_date, datetime.strptime('13:00', '%H:%M').time())
    all_periods = merged_periods + [(lunch_break_start, lunch_break_end)]
    
    # 合并所有时间段
    merged_periods = merge_datetime_periods(all_periods)
    
    # 剔除午休时间段
    merged_periods = [period for period in merged_periods 
                     if not (period[0] == lunch_break_start and period[1] == lunch_break_end)]

    # 计算总时长（排除午休时间后的实际时长）
    total_minutes = calculate_datetime_minutes(merged_periods)
    return min(total_minutes / 60, 9.0)

def get_absence_hours(name: str, day: str, xiujia_df: pd.DataFrame, waichu_df: pd.DataFrame, 
                     chuchai_df: pd.DataFrame) -> Tuple[float, float, str]:
    """获取员工某天的请假信息（原始时长、实际时长和请假描述）

    Args:
        name (str): 员工姓名
        day (str): 日期
        xiujia_df (pd.DataFrame): 休假单数据表
        waichu_df (pd.DataFrame): 外出单数据表
        chuchai_df (pd.DataFrame): 出差单数据表

    Returns:
        Tuple[float, float, str]: (原始请假时长, 实际请假时长, 请假描述)
    """
    # 准备数据表列表，每个元素为 (数据表, 类型列名, 开始时间列名, 结束时间列名)
    absence_dfs = [
        (xiujia_df, '休假类型', '开始时间', '结束时间'),
        (waichu_df, '类型', '开始时间', '结束时间'),
        (chuchai_df, '类型', '开始时间', '结束时间')
    ]

    # 获取原始请假时长和合并后的时间段
    raw_hours, merged_periods = get_raw_absence_hours(name, day, absence_dfs)

    # 获取实际请假时长（考虑午休时间）
    actual_hours = get_actual_absence_hours(merged_periods)

    # 获取请假描述
    description = get_absence_description(name, day, absence_dfs)

    return raw_hours, actual_hours, description

def analyze_attendance(kaoqin_df, xiujia_df, waichu_df, chuchai_df, employee_df, linshika_df, start_date, end_date, work_days, holiday_days):
    """分析考勤数据"""
    # 筛选时间范围内的考勤数据
    kaoqin_df['事件时间'] = pd.to_datetime(kaoqin_df['事件时间'], errors='coerce')
    kaoqin_df = kaoqin_df.dropna(subset=['事件时间'])
    
    date_range_mask = (kaoqin_df['事件时间'].dt.date >= start_date) & (kaoqin_df['事件时间'].dt.date <= end_date)
    kaoqin_df = kaoqin_df[date_range_mask]
    
    # 获取符合条件的员工名单和部门信息
    employee_info = employee_df[['姓名', '部门']].set_index('姓名').to_dict()['部门']
    qualified_employees = list(employee_info.keys())
    
    # 按天统计
    kaoqin_df['日期'] = kaoqin_df['事件时间'].dt.date
    all_dates = sorted(kaoqin_df['日期'].unique())
    
    # 只保留工作日
    work_dates = [date for date in all_dates if is_workday_from_calendar(date, work_days, holiday_days)]
    
    # 处理临时卡数据
    if not linshika_df.empty:
        linshika_df['借卡时间'] = pd.to_datetime(linshika_df['借卡时间']).dt.date
    
    # 统计表数据
    statistics_data = []
    # 汇总表数据
    summary_data = {
        '未打卡': {},
        '上/下班漏打卡': {},
        '出勤不足': {}
    }
    
    for day in work_dates:
        day_str = str(day)
        
        for name in qualified_employees:
            # 获取当天考勤记录
            day_df = kaoqin_df[(kaoqin_df['日期'] == day) & (kaoqin_df['持卡人员'] == name)]
            
            if not day_df.empty:
                times = day_df['事件时间'].sort_values()
                first_in = times.iloc[0]
                last_out = times.iloc[-1]
                work_hours = (last_out - first_in).total_seconds() / 3600
                first_in_str = first_in.strftime('%H:%M:%S')
                last_out_str = last_out.strftime('%H:%M:%S')
                record_count = len(day_df)
            else:
                work_hours = 0
                first_in_str = ''
                last_out_str = ''
                record_count = 0
            
            # 获取缺席申请时长
            raw_absence_hours = 0.0
            actual_absence_hours = 0.0
            absence_texts = []
            
            # 休假
            raw_h, actual_h, text = get_absence_hours(name, day_str, xiujia_df, waichu_df, chuchai_df)
            raw_absence_hours += raw_h
            actual_absence_hours += actual_h
            if text:
                absence_texts.append(text)
            
            # 应出勤时长 = 8.75 - 实际请假时长，最小为0
            required_hours = max(0, 8.75 - actual_absence_hours)
            
            # 判断状态
            status = '正常'
            if record_count == 0 and required_hours > 0:  # 修改未打卡判断逻辑
                status = '未打卡'
            elif record_count == 1 and work_hours < required_hours:
                status = '上/下班漏打卡'
            elif work_hours < required_hours:
                status = '出勤不足'
            elif work_hours >= required_hours:
                status = '正常'
            
            # 计算实际缺勤时间
            actual_absence_minutes = max(0, round((9 - actual_absence_hours - work_hours) * 60))
            
            # 检查是否有临时卡记录
            has_temp_card = False
            if not linshika_df.empty:
                temp_card_records = linshika_df[
                    (linshika_df['姓名'] == name) & 
                    (linshika_df['借卡时间'] == day)
                ]
                has_temp_card = not temp_card_records.empty
            
            # 添加到统计表
            statistics_data.append({
                '日期': day_str,
                '姓名': name,
                '部门': employee_info.get(name, ''),
                '上班时间': first_in_str,
                '下班时间': last_out_str,
                '工作时长': round(work_hours, 2),
                '当天请假时长': round(raw_absence_hours, 2),
                '午休时长': 1.0,  # 固定1小时
                '实际请假时长': round(actual_absence_hours, 2),
                '实际缺勤时间[分钟]': actual_absence_minutes,
                '应出勤时长': round(required_hours, 2),
                '情况说明': '; '.join(absence_texts) if absence_texts else '',
                '状态': status,
                '临时卡': '是' if has_temp_card else '否',
                '备注': ''
            })
            
            # 基于统计表结果生成异常汇总表数据
            if status in ['未打卡', '上/下班漏打卡', '出勤不足']:
                if name not in summary_data[status]:
                    summary_data[status][name] = 0
                summary_data[status][name] += 1
    
    return statistics_data, summary_data

def get_cell_color(header: str, value: str, row_data: dict) -> Optional[str]:
    """根据单元格内容确定背景色
    
    Args:
        header: 列标题
        value: 单元格值
        row_data: 行数据字典
        
    Returns:
        Optional[str]: 背景色代码，如果不需要背景色则返回None
    """
    if header == '状态':
        return get_status_color(row_data['状态'])
    elif header == '实际缺勤时间[分钟]':
        # 如果实际缺勤时间 > 0 且状态是正常，标记为莫兰迪色系
        minutes = int(value)
        if minutes > 0 and row_data['状态'] == '正常':
            return 'D4E4BC'  # 莫兰迪淡绿色
    elif header == '临时卡' and value == '是':
        return '92D050'  # 绿色
    elif header in ['未打卡次数', '上/下班漏打卡次数', '出勤不足次数']:
        count = int(value)
        if count > 3:
            return 'E6B8AF'  # 莫兰迪红
        elif count == 3:
            return 'FFE599'  # 莫兰迪黄
    return None

def save_reports(statistics_data, summary_data, start_date, end_date):
    """保存统计表和汇总表到同一个Excel文件的不同sheet中"""
    # 检查是否有任何数据需要保存
    if not statistics_data and not (summary_data and any(summary_data.values())):
        print(f"没有考勤数据，跳过生成报表")
        return
    
    # 格式化日期字符串
    start_date_str = start_date.strftime('%Y年%m月%d日')
    end_date_str = end_date.strftime('%Y年%m月%d日')
    filename = f"考勤报表{start_date_str}-{end_date_str}.xlsx"
    
    # 创建一个新的Excel工作簿
    wb = Workbook()
    
    # 保存统计表
    if statistics_data:
        # 删除默认sheet
        if wb.active:
            wb.remove(wb.active)
            
        # 创建统计表sheet
        ws_stats = wb.create_sheet(title='考勤统计')
        
        # 创建DataFrame并写入数据
        df_stats = pd.DataFrame(statistics_data)
        
        # 写入表头
        headers = list(df_stats.columns)
        for col, header in enumerate(headers, 1):
            cell = ws_stats.cell(row=1, column=col)
            apply_cell_format(cell, header, is_header=True, bg_color='CCCCCC', bold=True)
        
        # 设置列宽 - 考勤统计表
        # 日期(15) 姓名(12) 部门(25) 上班时间(12) 下班时间(12) 工作时长(10) 
        # 当天请假时长(12) 午休时长(10) 实际请假时长(12) 实际缺勤时间[分钟](18) 
        # 应出勤时长(12) 情况说明(50) 状态(20) 临时卡(8) 备注(20)
        set_column_widths(ws_stats, [15, 12, 25, 12, 12, 10, 12, 10, 12, 18, 12, 50, 20, 8, 20])
        
        # 写入数据
        for row_idx, row_data in enumerate(statistics_data, 2):
            for col_idx, header in enumerate(headers, 1):
                cell = ws_stats.cell(row=row_idx, column=col_idx)
                value = str(row_data[header])
                bg_color = get_cell_color(header, value, row_data)
                apply_cell_format(cell, value, bg_color=bg_color if bg_color else '')
            
        # 设置行高
        set_row_heights(ws_stats, 2, len(statistics_data) + 1, 25)
    
    # 保存汇总表
    if summary_data and any(summary_data.values()):
        # 获取所有员工姓名
        all_employees = set()
        for category_data in summary_data.values():
            all_employees.update(category_data.keys())
        
        all_employees = sorted(list(all_employees))
        
        # 创建汇总数据
        summary_rows = []
        for employee in all_employees:
            row = {
                '员工姓名': employee,
                '未打卡次数': summary_data['未打卡'].get(employee, 0),
                '上/下班漏打卡次数': summary_data['上/下班漏打卡'].get(employee, 0),
                '出勤不足次数': summary_data['出勤不足'].get(employee, 0),
                '总异常次数': 0,
                '备注': ''
            }
            row['总异常次数'] = row['未打卡次数'] + row['上/下班漏打卡次数'] + row['出勤不足次数']
            summary_rows.append(row)
        
        # 按总异常次数排序
        summary_rows.sort(key=lambda x: x['总异常次数'], reverse=True)
        
        if summary_rows:
            # 如果没有统计表，需要删除默认sheet
            if not statistics_data and wb.active:
                wb.remove(wb.active)
                
            # 创建汇总表sheet
            ws_summary = wb.create_sheet(title='异常汇总')
            
            # 添加标题行
            title = f"异常考勤汇总表 ({start_date_str} - {end_date_str})"
            title_cell = ws_summary.cell(row=1, column=1)
            apply_cell_format(title_cell, title, is_header=True, bg_color='D0E0E3',  # 莫兰迪蓝灰
                            bold=True, font_size=14)
            ws_summary.merge_cells('A1:F1')
            
            # 写入表头
            headers = ['员工姓名', '未打卡次数', '上/下班漏打卡次数', '出勤不足次数', '总异常次数', '备注']
            for col, header in enumerate(headers, 1):
                cell = ws_summary.cell(row=2, column=col)
                apply_cell_format(cell, header, is_header=True, bg_color='B4C7E7',  # 莫兰迪蓝
                                bold=True, font_color='000000', font_size=12)
            
            # 写入数据
            for row_idx, row_data in enumerate(summary_rows, 3):
                for col_idx, header in enumerate(headers, 1):
                    cell = ws_summary.cell(row=row_idx, column=col_idx)
                    value = str(row_data[header])
                    bg_color = get_cell_color(header, value, row_data)
                    apply_cell_format(cell, value, bg_color=bg_color if bg_color else '')
            
            # 设置列宽 - 异常汇总表
            # 员工姓名(15) 未打卡次数(12) 上/下班漏打卡次数(22) 出勤不足次数(12) 总异常次数(12) 备注(30)
            set_column_widths(ws_summary, [15, 12, 22, 12, 12, 30])
            set_row_heights(ws_summary, 1, 1, 35)  # 标题行
            set_row_heights(ws_summary, 2, 2, 30)  # 表头行
            set_row_heights(ws_summary, 3, len(summary_rows) + 2, 30)  # 数据行
    
    # 保存Excel文件
    try:
        wb.save(filename)
        print(f"报表已保存: {filename}")
    except Exception as e:
        print(f"保存Excel文件时出错: {e}")

def read_calendar_file(file_path):
    """读取日历文件，解析法定节假日和调休工作日信息"""
    try:
        # 读取Excel文件的所有sheet
        excel_file = pd.ExcelFile(file_path)
        available_sheets = excel_file.sheet_names
        
        # 存储所有工作日信息
        work_days = set()
        holiday_days = set()
        
        # 查找目标sheet
        target_sheet = '法定节假日和调休工作日'
        if target_sheet in available_sheets:
            print(f"读取日历文件: {target_sheet}")
            df = pd.read_excel(file_path, sheet_name=target_sheet, engine='openpyxl')
            
            # 检查必要的列是否存在
            if '日期' not in df.columns or '日期类型' not in df.columns:
                print(f"警告: {target_sheet} 缺少必要的列（日期 或 日期类型）")
                return set(), set()
            
            # 解析日期类型
            for _, row in df.iterrows():
                date_str = str(row['日期']).strip()
                date_type = str(row['日期类型']).strip()
                
                # 跳过空行或无效数据
                if pd.isna(date_str) or date_str == 'nan' or date_str == '':
                    continue
                
                # 标准化日期格式
                try:
                    if ' ' in date_str:
                        date_str = date_str.split(' ')[0]  # 只取日期部分
                    
                    # 验证日期格式
                    datetime.strptime(date_str, '%Y-%m-%d')
                    
                    if date_type == '调休工作日' or date_type == '工作日':
                        work_days.add(date_str)
                        if date_type == '调休工作日':
                            print(f"  调休工作日: {date_str}")
                        else:
                            print(f"  工作日: {date_str}")
                    elif date_type == '法定节假日':
                        holiday_days.add(date_str)
                        print(f"  法定节假日: {date_str}")
                    elif date_type == '周末休息':
                        # 周末不参与统计，跳过
                        pass
                    else:
                        print(f"  未知日期类型: {date_str} - {date_type}")
                        
                except ValueError as e:
                    print(f"  日期格式错误: {date_str} - {e}")
                    continue
        else:
            print(f"错误: 找不到目标sheet '{target_sheet}'")
            print(f"可用的sheet: {available_sheets}")
            return set(), set()
        
        print(f"日历解析完成: 工作日 {len(work_days)} 天, 法定节假日 {len(holiday_days)} 天")
        return work_days, holiday_days
        
    except Exception as e:
        print(f"读取日历文件出错: {e}")
        return set(), set()

def is_workday_from_calendar(date_obj, work_days, holiday_days):
    """根据日历文件判断是否为工作日"""
    date_str = date_obj.strftime('%Y-%m-%d')
    
    # 如果是法定节假日，不是工作日
    if date_str in holiday_days:
        return False
    
    # 如果是调休工作日或普通工作日，是工作日
    if date_str in work_days:
        return True
    
    # 其他情况按原来的逻辑判断（周一到周五是工作日）
    return date_obj.weekday() < 5

def main():
    """主函数"""
    print("=" * 50)
    print("考勤统计分析程序")
    print("=" * 50)
    print("\n开始考勤分析...")
    input("按回车键开始分析...")  # 添加暂停

    # 读取员工花名册
    print("\n[1/6] 读取员工花名册...")
    employee_df, error = read_excel_file('员工花名册.xlsx', '员工花名册')
    if error:
        print(error)
        input("\n处理失败，按回车键退出...")
        return
    
    # 筛选符合条件的员工
    qualified_employees = filter_employees(employee_df)
    print(f"筛选出符合条件的员工: {len(qualified_employees)} 人")
    
    # 读取考勤数据
    print("\n[2/6] 读取考勤数据...")
    kaoqin_df, error = read_excel_file('原始数据.xlsx', '原始数据')
    if error:
        print(error)
        input("\n处理失败，按回车键退出...")
        return
    
    # 读取休假数据
    print("\n[3/6] 读取休假数据...")
    xiujia_df, error = read_excel_file('休假单.xlsx', '休假单')
    if error:
        print(error)
        input("\n处理失败，按回车键退出...")
        return
    
    # 读取外出数据
    print("\n[4/6] 读取外出数据...")
    waichu_df, error = read_excel_file('外出单.xlsx', '外出单')
    if error:
        print(error)
        input("\n处理失败，按回车键退出...")
        return
    
    # 读取出差数据
    print("\n[5/6] 读取出差数据...")
    chuchai_df, error = read_excel_file('出差单.xlsx', '出差单')
    if error:
        print(error)
        input("\n处理失败，按回车键退出...")
        return
    
    # 读取临时卡数据
    print("\n[6/6] 读取临时卡数据...")
    linshika_df, error = read_excel_file('临时卡.xlsx', '临时卡')
    if error:
        print(f"警告: {error}")
        print("继续处理，临时卡数据将被忽略...")
        linshika_df = pd.DataFrame()  # 使用空DataFrame继续处理
    
    # 读取日历文件
    print("\n[进行中] 读取日历文件...")
    work_days, holiday_days = read_calendar_file('日历.xlsx')
    
    # 获取考勤数据的时间范围
    print("\n[进行中] 分析考勤数据时间范围...")
    kaoqin_df['事件时间'] = pd.to_datetime(kaoqin_df['事件时间'], errors='coerce')
    kaoqin_df = kaoqin_df.dropna(subset=['事件时间'])
    
    if kaoqin_df.empty:
        print("错误: 没有有效的考勤数据")
        input("\n处理失败，按回车键退出...")
        return
        
    start_date = kaoqin_df['事件时间'].dt.date.min()
    end_date = kaoqin_df['事件时间'].dt.date.max()
    print(f"考勤数据时间范围: {start_date} 到 {end_date}")
    
    # 分析考勤数据
    print("\n[进行中] 分析考勤数据...")
    statistics_data, summary_data = analyze_attendance(
        kaoqin_df, xiujia_df, waichu_df, chuchai_df, 
        qualified_employees, linshika_df, start_date, end_date, work_days, holiday_days
    )
    
    # 保存报表
    print("\n[进行中] 生成考勤报表...")
    save_reports(statistics_data, summary_data, start_date, end_date)
    
    print("\n考勤分析完成！")
    print("\n生成的文件：")
    print("1. 考勤报表（包含考勤统计和异常汇总）")
    print("\n注意：请检查生成的文件是否完整。")
    input("\n处理完成，按回车键退出...")

if __name__ == "__main__":
    main()
